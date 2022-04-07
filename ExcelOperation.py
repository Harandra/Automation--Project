import openpyxl


class ExcelUtil():
    def loadfile(self, path):
        self.path = path
        self.wb_obj = openpyxl.load_workbook(path)
        self.sheet = self.wb_obj.active

    def rowCount(self):
        return self.sheet.min_row

    def readData(self, row, col):
        return self.sheet.cell(row=row, column=col).value

    def writeDate(self, row, col, val):
        self.sheet.cell(row=row, column=col).value = val
        self.wb_obj.save(self.path)
